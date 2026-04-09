"""提现服务"""

import uuid
from typing import Optional, List, Dict, Any
from datetime import datetime
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..database import get_db_manager, Withdrawal, Account, User
from ..database.manager import DatabaseManager


class WithdrawalService:
    """提现服务（线下打款模式）"""

    def __init__(self):
        self.db_manager: DatabaseManager = get_db_manager()

    async def create_withdrawal(
        self,
        user_id: str,
        amount: int,
        method: str,
        account: str,
        name: str,
        note: str = None
    ) -> Dict[str, Any]:
        """
        创建提现申请

        Args:
            user_id: 用户ID
            amount: 提现金额（分）
            method: 提现方式 (wechat, alipay, bank)
            account: 提现账号
            name: 收款人姓名
            note: 用户备注

        Returns:
            提现记录信息
        """
        # 验证提现方式
        if method not in ["wechat", "alipay", "bank"]:
            raise ValueError("无效的提现方式")

        # 获取用户账户
        user_account = await self.db_manager.get_account_by_user(user_id)
        if not user_account:
            raise ValueError("用户账户不存在")

        # 检查余额是否充足
        if user_account.balance < amount:
            raise ValueError(f"余额不足，当前余额: {user_account.balance / 100:.2f}元")

        # 生成提现单号
        withdrawal_id = f"WD{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"

        # 创建提现记录
        withdrawal = Withdrawal(
            withdrawal_id=withdrawal_id,
            user_id=user_id,
            amount=amount,
            withdrawal_method=method,
            withdrawal_account=account,
            withdrawal_name=name,
            status="pending",
            user_note=note
        )

        withdrawal = await self.db_manager.create_withdrawal(withdrawal)

        logger.info(f"用户 {user_id} 创建提现申请: {withdrawal_id}, 金额: {amount / 100:.2f}元")

        return {
            "withdrawal_id": withdrawal.withdrawal_id,
            "amount": withdrawal.amount,
            "amount_yuan": withdrawal.amount / 100,
            "status": withdrawal.status,
            "created_at": withdrawal.created_at.isoformat() if withdrawal.created_at else None
        }

    async def get_user_withdrawals(
        self,
        user_id: str,
        limit: int = 20,
        offset: int = 0
    ) -> List[Dict[str, Any]]:
        """获取用户提现记录"""
        withdrawals = await self.db_manager.get_user_withdrawals(user_id, limit, offset)

        result = []
        for w in withdrawals:
            result.append({
                "id": w.id,
                "withdrawal_id": w.withdrawal_id,
                "amount": w.amount,
                "amount_yuan": w.amount / 100,
                "withdrawal_method": w.withdrawal_method,
                "withdrawal_account": self.mask_account(w.withdrawal_account, w.withdrawal_method),
                "withdrawal_name": w.withdrawal_name,
                "status": w.status,
                "user_note": w.user_note,
                "review_note": w.review_note,
                "created_at": w.created_at.isoformat() if w.created_at else None,
                "reviewed_at": w.reviewed_at.isoformat() if w.reviewed_at else None,
                "completed_at": w.completed_at.isoformat() if w.completed_at else None
            })

        return result

    async def get_withdrawals_for_admin(
        self,
        status: str = None,
        limit: int = 20,
        offset: int = 0
    ) -> List[Dict[str, Any]]:
        """获取提现记录（管理员用）"""
        withdrawals = await self.db_manager.get_all_withdrawals(status, limit, offset)

        result = []
        for w in withdrawals:
            # 获取用户信息
            user = await self.db_manager.get_user_by_id(w.user_id)
            result.append({
                "id": w.id,
                "withdrawal_id": w.withdrawal_id,
                "user_id": w.user_id,
                "username": user.username if user else "",
                "phone": user.phone if user else "",
                "amount": w.amount,
                "amount_yuan": w.amount / 100,
                "withdrawal_method": w.withdrawal_method,
                "withdrawal_account": w.withdrawal_account,
                "withdrawal_name": w.withdrawal_name,
                "status": w.status,
                "user_note": w.user_note,
                "admin_id": w.admin_id,
                "review_note": w.review_note,
                "payment_proof": w.payment_proof,
                "created_at": w.created_at.isoformat() if w.created_at else None,
                "reviewed_at": w.reviewed_at.isoformat() if w.reviewed_at else None,
                "completed_at": w.completed_at.isoformat() if w.completed_at else None
            })

        return result

    async def approve_withdrawal(
        self,
        withdrawal_id: str,
        admin_id: str,
        note: str = None
    ) -> Dict[str, Any]:
        """
        审核通过提现申请

        审核通过后，扣除用户余额，状态变为 approved
        """
        withdrawal = await self.db_manager.get_withdrawal_by_id(withdrawal_id)
        if not withdrawal:
            raise ValueError("提现记录不存在")

        if withdrawal.status != "pending":
            raise ValueError(f"提现申请状态不是待审核，当前状态: {withdrawal.status}")

        # 获取用户账户
        user_account = await self.db_manager.get_account_by_user(withdrawal.user_id)
        if not user_account:
            raise ValueError("用户账户不存在")

        # 再次检查余额
        if user_account.balance < withdrawal.amount:
            raise ValueError("用户余额不足，无法通过审核")

        # 扣除余额
        old_balance = user_account.balance
        user_account.balance -= withdrawal.amount
        await self.db_manager.update_account(user_account)

        # 记录交易
        await self.db_manager.add_transaction(
            user_id=withdrawal.user_id,
            transaction_type="withdrawal",
            amount=-withdrawal.amount,
            description=f"提现申请: {withdrawal_id}",
            apply_account_change=False,
            balance_after=user_account.balance,
            points_after=user_account.points,
        )

        # 更新提现记录
        withdrawal.status = "approved"
        withdrawal.admin_id = admin_id
        withdrawal.review_note = note
        withdrawal.reviewed_at = datetime.utcnow()
        await self.db_manager.update_withdrawal(withdrawal)

        logger.info(f"管理员 {admin_id} 审核通过提现: {withdrawal_id}, 扣除余额: {withdrawal.amount / 100:.2f}元")

        return {
            "withdrawal_id": withdrawal.withdrawal_id,
            "status": withdrawal.status,
            "old_balance": old_balance,
            "new_balance": user_account.balance
        }

    async def reject_withdrawal(
        self,
        withdrawal_id: str,
        admin_id: str,
        reason: str
    ) -> Dict[str, Any]:
        """
        审核拒绝提现申请

        审核拒绝后，余额不变，状态变为 rejected
        """
        withdrawal = await self.db_manager.get_withdrawal_by_id(withdrawal_id)
        if not withdrawal:
            raise ValueError("提现记录不存在")

        if withdrawal.status != "pending":
            raise ValueError(f"提现申请状态不是待审核，当前状态: {withdrawal.status}")

        # 更新提现记录
        withdrawal.status = "rejected"
        withdrawal.admin_id = admin_id
        withdrawal.review_note = reason
        withdrawal.reviewed_at = datetime.utcnow()
        await self.db_manager.update_withdrawal(withdrawal)

        logger.info(f"管理员 {admin_id} 审核拒绝提现: {withdrawal_id}, 原因: {reason}")

        return {
            "withdrawal_id": withdrawal.withdrawal_id,
            "status": withdrawal.status,
            "reason": reason
        }

    async def mark_as_paid(
        self,
        withdrawal_id: str,
        admin_id: str,
        payment_proof: str = None
    ) -> Dict[str, Any]:
        """
        标记提现为已打款

        线下打款完成后调用，状态变为 completed
        """
        withdrawal = await self.db_manager.get_withdrawal_by_id(withdrawal_id)
        if not withdrawal:
            raise ValueError("提现记录不存在")

        if withdrawal.status != "approved":
            raise ValueError(f"提现申请状态不是已通过，当前状态: {withdrawal.status}")

        # 更新提现记录
        withdrawal.status = "completed"
        withdrawal.payment_proof = payment_proof
        withdrawal.completed_at = datetime.utcnow()
        await self.db_manager.update_withdrawal(withdrawal)

        logger.info(f"管理员 {admin_id} 标记提现为已打款: {withdrawal_id}")

        return {
            "withdrawal_id": withdrawal.withdrawal_id,
            "status": withdrawal.status
        }

    async def export_approved_withdrawals_to_excel(
        self,
        start_date: str = None,
        end_date: str = None
    ) -> bytes:
        """
        导出待打款订单到Excel

        导出已审核通过但未打款的订单
        """
        # 获取待打款订单
        withdrawals = await self.db_manager.get_all_withdrawals(status="approved", limit=10000, offset=0)

        # 过滤日期范围
        if start_date or end_date:
            filtered = []
            for w in withdrawals:
                if w.created_at:
                    if start_date and w.created_at < datetime.fromisoformat(start_date):
                        continue
                    if end_date and w.created_at > datetime.fromisoformat(end_date):
                        continue
                filtered.append(w)
            withdrawals = filtered

        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "待打款订单"

        # 设置表头
        headers = [
            "提现单号", "用户ID", "用户手机号", "收款人姓名",
            "收款账号", "提现方式", "金额(元)", "申请时间", "审核时间"
        ]

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        withdrawal_method_map = {
            "wechat": "微信",
            "alipay": "支付宝",
            "bank": "银行卡"
        }

        for row, withdrawal in enumerate(withdrawals, start=2):
            user = await self.db_manager.get_user_by_id(withdrawal.user_id)

            ws.cell(row=row, column=1, value=withdrawal.withdrawal_id)
            ws.cell(row=row, column=2, value=withdrawal.user_id)
            ws.cell(row=row, column=3, value=user.phone if user else "")
            ws.cell(row=row, column=4, value=withdrawal.withdrawal_name or "")
            ws.cell(row=row, column=5, value=withdrawal.withdrawal_account or "")
            ws.cell(row=row, column=6, value=withdrawal_method_map.get(withdrawal.withdrawal_method, ""))
            ws.cell(row=row, column=7, value=withdrawal.amount / 100)
            ws.cell(row=row, column=8, value=withdrawal.created_at.strftime("%Y-%m-%d %H:%M:%S") if withdrawal.created_at else "")
            ws.cell(row=row, column=9, value=withdrawal.reviewed_at.strftime("%Y-%m-%d %H:%M:%S") if withdrawal.reviewed_at else "")

        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # 保存到字节流
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"导出待打款订单: {len(withdrawals)} 条记录")

        return output.read()

    async def cancel_withdrawal(self, withdrawal_id: str, user_id: str) -> Dict[str, Any]:
        """
        用户取消待审核的提现申请

        只有待审核状态才能取消
        """
        withdrawal = await self.db_manager.get_withdrawal_by_id(withdrawal_id)
        if not withdrawal:
            raise ValueError("提现记录不存在")

        if withdrawal.user_id != user_id:
            raise ValueError("无权取消此提现申请")

        if withdrawal.status != "pending":
            raise ValueError(f"只有待审核状态才能取消，当前状态: {withdrawal.status}")

        # 更新状态为已取消（使用 rejected 表示用户主动取消）
        withdrawal.status = "rejected"
        withdrawal.review_note = "用户主动取消"
        await self.db_manager.update_withdrawal(withdrawal)

        logger.info(f"用户 {user_id} 取消提现申请: {withdrawal_id}")

        return {
            "withdrawal_id": withdrawal.withdrawal_id,
            "status": withdrawal.status
        }

    @staticmethod
    def mask_account(account: str, method: str) -> str:
        """脱敏账号"""
        if not account:
            return ""

        if method == "bank":
            # 银行卡显示前4后4
            if len(account) > 8:
                return account[:4] + " **** **** " + account[-4:]
            return account
        else:
            # 微信/支付宝显示前2后2
            if len(account) > 4:
                return account[:2] + "***" + account[-2:]
            return account


# 全局实例
_withdrawal_service: Optional[WithdrawalService] = None


def get_withdrawal_service() -> WithdrawalService:
    """获取提现服务实例"""
    global _withdrawal_service
    if _withdrawal_service is None:
        _withdrawal_service = WithdrawalService()
    return _withdrawal_service
